import csv
import io
import json
import hashlib
import math
import os
import re
import secrets
import sqlite3
import unicodedata
from datetime import datetime, timedelta, timezone

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from flask import Flask, jsonify, request, send_file, send_from_directory

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "flashcards.db")

app = Flask(__name__)

AI_CONNECT_TIMEOUT = 10
AI_READ_TIMEOUT = 60
AI_RETRY_COUNT = 2
AI_RETRY_BACKOFF = 0.6

_ai_retry = Retry(
    total=AI_RETRY_COUNT,
    backoff_factor=AI_RETRY_BACKOFF,
    status_forcelist=(429, 500, 502, 503, 504),
    allowed_methods=frozenset(["POST"]),
    raise_on_status=False,
)
_ai_session = requests.Session()
_ai_session.mount("https://", HTTPAdapter(max_retries=_ai_retry))


def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def init_db():
    with get_db() as conn:
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS cards (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                collection_id INTEGER,
                question TEXT NOT NULL,
                answer TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        cols = [row["name"] for row in conn.execute("PRAGMA table_info(cards)").fetchall()]
        if "collection_id" not in cols:
            conn.execute("ALTER TABLE cards ADD COLUMN collection_id INTEGER")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS collections (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS study_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                collection_id INTEGER,
                created_at TEXT NOT NULL,
                duration_sec INTEGER NOT NULL,
                total INTEGER NOT NULL,
                correct INTEGER NOT NULL,
                incorrect INTEGER NOT NULL,
                easy INTEGER NOT NULL,
                hard INTEGER NOT NULL,
                details_json TEXT NOT NULL
            )
            """
        )
        cols = [row["name"] for row in conn.execute("PRAGMA table_info(study_sessions)").fetchall()]
        if "collection_id" not in cols:
            conn.execute("ALTER TABLE study_sessions ADD COLUMN collection_id INTEGER")
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS card_difficulty (
                card_id INTEGER PRIMARY KEY,
                difficulty TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS study_goals (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                collection_id INTEGER NOT NULL UNIQUE,
                days_json TEXT NOT NULL,
                created_at TEXT NOT NULL,
                updated_at TEXT NOT NULL
            )
            """
        )
        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS share_links (
                token TEXT PRIMARY KEY,
                type TEXT NOT NULL,
                name TEXT,
                payload_json TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
            """
        )
        cols = [row["name"] for row in conn.execute("PRAGMA table_info(share_links)").fetchall()]
        if "expires_at" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN expires_at TEXT")
        if "max_uses" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN max_uses INTEGER")
        if "uses" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN uses INTEGER DEFAULT 0")
        if "password_hash" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN password_hash TEXT")
        if "disabled" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN disabled INTEGER DEFAULT 0")
        if "manage_token" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN manage_token TEXT")
        if "session_token" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN session_token TEXT")
        if "collection_id" not in cols:
            conn.execute("ALTER TABLE share_links ADD COLUMN collection_id INTEGER")

        conn.execute(
            """
            CREATE TABLE IF NOT EXISTS exam_sessions (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                collection_id INTEGER,
                topic TEXT,
                name TEXT,
                created_at TEXT NOT NULL,
                started_at TEXT,
                duration_sec INTEGER NOT NULL,
                time_limit_sec INTEGER,
                allow_multi INTEGER NOT NULL,
                total INTEGER NOT NULL,
                answered INTEGER NOT NULL,
                correct INTEGER NOT NULL,
                incorrect INTEGER NOT NULL,
                details_json TEXT NOT NULL
            )
            """
        )
        cols = [row["name"] for row in conn.execute("PRAGMA table_info(exam_sessions)").fetchall()]
        if "collection_id" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN collection_id INTEGER")
        if "topic" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN topic TEXT")
        if "name" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN name TEXT")
        if "created_at" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN created_at TEXT")
        if "started_at" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN started_at TEXT")
        if "duration_sec" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN duration_sec INTEGER")
        if "time_limit_sec" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN time_limit_sec INTEGER")
        if "allow_multi" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN allow_multi INTEGER DEFAULT 0")
        if "total" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN total INTEGER DEFAULT 0")
        if "answered" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN answered INTEGER DEFAULT 0")
        if "correct" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN correct INTEGER DEFAULT 0")
        if "incorrect" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN incorrect INTEGER DEFAULT 0")
        if "details_json" not in cols:
            conn.execute("ALTER TABLE exam_sessions ADD COLUMN details_json TEXT DEFAULT '{}'")

        conn.execute("CREATE INDEX IF NOT EXISTS idx_cards_collection_id ON cards(collection_id)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_cards_created_at ON cards(created_at)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_sessions_collection_id ON study_sessions(collection_id)"
        )
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_study_sessions_created_at ON study_sessions(created_at)"
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_card_difficulty_difficulty ON card_difficulty(difficulty)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_share_links_manage_token ON share_links(manage_token)")
        conn.execute("CREATE INDEX IF NOT EXISTS idx_share_links_collection_id ON share_links(collection_id)")
        conn.execute(
            "CREATE INDEX IF NOT EXISTS idx_exam_sessions_collection_id ON exam_sessions(collection_id)"
        )
        conn.execute("CREATE INDEX IF NOT EXISTS idx_exam_sessions_created_at ON exam_sessions(created_at)")


def row_to_dict(row):
    return {
        "id": row["id"],
        "collection_id": row["collection_id"],
        "question": row["question"],
        "answer": row["answer"],
        "created_at": row["created_at"],
    }


def extract_output_text(response_json):
    # Best-effort extraction for Responses API output text
    output_texts = []
    for item in response_json.get("output", []):
        for content in item.get("content", []):
            if content.get("type") in ("output_text", "text"):
                output_texts.append(content.get("text", ""))
    return "\n".join(output_texts).strip()


def extract_gemini_text(response_json):
    candidates = response_json.get("candidates") or []
    for cand in candidates:
        content = cand.get("content") or {}
        parts = content.get("parts") or []
        for part in parts:
            text = part.get("text")
            if text:
                return text
    return ""


def detect_provider(api_key, hint=""):
    hint = (hint or "").strip().lower()
    if hint in ("openai", "gemini"):
        return hint
    if api_key.startswith("AIza"):
        return "gemini"
    return "openai"


def api_error(message, status=400, code=None, details=None):
    payload = {"error": message}
    if code:
        payload["code"] = code
    if details is not None:
        payload["details"] = details
    return jsonify(payload), status


def _get_request_json(required=True):
    data = request.get_json(silent=True)
    if data is None:
        if required:
            return None, api_error("JSON inválido.", 400, code="invalid_json")
        return {}, None
    if not isinstance(data, dict):
        return None, api_error("JSON inválido.", 400, code="invalid_json")
    return data, None


def _parse_int(value, label, required=False, min_value=None, max_value=None):
    if value is None or (isinstance(value, str) and not value.strip()):
        if required:
            return None, api_error(f"{label} é obrigatório.", 400, code="required")
        return None, None
    try:
        parsed = int(value)
    except (TypeError, ValueError):
        return None, api_error(f"{label} inválido.", 400, code="invalid")
    if min_value is not None and parsed < min_value:
        return None, api_error(f"{label} inválido.", 400, code="invalid")
    if max_value is not None and parsed > max_value:
        return None, api_error(f"{label} inválido.", 400, code="invalid")
    return parsed, None


def _trim_error_details(text, limit=500):
    if text is None:
        return None
    text = str(text)
    if len(text) <= limit:
        return text
    return text[:limit] + "..."


def _coerce_bool(value):
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return bool(value)
    if isinstance(value, str):
        raw = value.strip().lower()
        if raw in ("1", "true", "yes", "y", "on", "sim"):
            return True
        if raw in ("0", "false", "no", "n", "off", "nao", "não"):
            return False
    return None


def _normalize_mcq_correct(raw):
    if raw is None:
        return []
    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        raw = [raw]
    elif isinstance(raw, str):
        stripped = raw.strip()
        if not stripped:
            return []
        raw = [part for part in stripped.replace(";", ",").replace("|", ",").split(",") if part.strip()]
    elif not isinstance(raw, list):
        raw = [raw]

    indices = []
    for item in raw:
        if item is None or isinstance(item, bool):
            continue
        if isinstance(item, (int, float)) and not isinstance(item, bool):
            try:
                idx = int(item)
            except (TypeError, ValueError):
                continue
            if 1 <= idx <= 4:
                idx -= 1
            if 0 <= idx <= 3:
                indices.append(idx)
            continue

        if isinstance(item, str):
            s = item.strip().lower()
            if not s:
                continue
            if s.isdigit():
                idx = int(s)
                if 1 <= idx <= 4:
                    idx -= 1
                if 0 <= idx <= 3:
                    indices.append(idx)
                continue
            first = s[0]
            if first in ("a", "b", "c", "d"):
                indices.append({"a": 0, "b": 1, "c": 2, "d": 3}[first])

    seen = set()
    normalized = []
    for idx in indices:
        if idx in seen:
            continue
        seen.add(idx)
        normalized.append(idx)
    return normalized


_MCQ_EXPLANATION_STOPWORDS = {
    "a",
    "an",
    "and",
    "as",
    "aws",
    "amazon",
    "cloud",
    "da",
    "das",
    "de",
    "do",
    "dos",
    "e",
    "em",
    "na",
    "nas",
    "no",
    "nos",
    "o",
    "os",
    "ou",
    "para",
    "por",
    "service",
    "services",
    "servico",
    "servicos",
    "serviço",
    "serviços",
    "the",
    "um",
    "uma",
}


def _normalize_text_match(text):
    raw = str(text or "")
    raw = unicodedata.normalize("NFKD", raw)
    raw = "".join(ch for ch in raw if not unicodedata.combining(ch))
    raw = raw.lower()
    raw = re.sub(r"[^a-z0-9]+", " ", raw)
    return " ".join(raw.split())


def _score_options_in_explanation(options, explanation):
    exp_norm = _normalize_text_match(explanation)
    exp_pad = f" {exp_norm} "
    scores = []
    for opt in options:
        opt_norm = _normalize_text_match(opt)
        score = 0
        if opt_norm:
            if f" {opt_norm} " in exp_pad:
                score += 3
            elif opt_norm in exp_norm:
                score += 2
        tokens = []
        for tok in opt_norm.split():
            if tok in _MCQ_EXPLANATION_STOPWORDS:
                continue
            if any(ch.isdigit() for ch in tok) or len(tok) >= 5:
                tokens.append(tok)
        for tok in set(tokens):
            if f" {tok} " in exp_pad:
                score += 1
        scores.append(score)
    return scores


def _fix_mcq_correct_from_explanation(question):
    if not isinstance(question, dict):
        return question
    options = question.get("options") or []
    correct = question.get("correct") or []
    explanation = str(question.get("explanation") or "").strip()
    if not isinstance(options, list) or len(options) != 4:
        return question
    if not isinstance(correct, list) or len(correct) != 1:
        return question
    if not explanation:
        return question
    try:
        current = int(correct[0])
    except (TypeError, ValueError):
        return question
    if current < 0 or current > 3:
        return question

    scores = _score_options_in_explanation(options, explanation)
    if not scores or len(scores) != 4:
        return question
    best_score = max(scores)
    if best_score < 2:
        return question
    best_indices = [i for i, s in enumerate(scores) if s == best_score]
    if len(best_indices) != 1:
        return question
    best_idx = best_indices[0]
    if best_idx == current:
        return question
    if scores[current] >= 2:
        return question

    fixed = dict(question)
    fixed["correct"] = [best_idx]
    return fixed


@app.route("/")
def index():
    return send_from_directory("templates", "index.html")


@app.route("/study")
def study():
    return send_from_directory("templates", "study.html")


@app.route("/exam")
def exam():
    return send_from_directory("templates", "exam.html")


@app.route("/share/<string:token>")
def share_page(token):
    return send_from_directory("templates", "share.html")


@app.route("/challenge/<string:token>")
def challenge_page(token):
    return send_from_directory("templates", "challenge.html")


@app.route("/api/cards", methods=["GET"])
def list_cards():
    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None
    with get_db() as conn:
        if collection_id_val is not None:
            if not _collection_exists(conn, collection_id_val):
                return api_error("Coleção não encontrada.", 404, code="collection_not_found")
            rows = conn.execute(
                """
                SELECT id, collection_id, question, answer, created_at
                FROM cards
                WHERE collection_id = ?
                ORDER BY id DESC
                """,
                (collection_id_val,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, collection_id, question, answer, created_at FROM cards ORDER BY id DESC"
            ).fetchall()
    return jsonify([row_to_dict(row) for row in rows])


@app.route("/api/study/cards", methods=["GET"])
def study_cards():
    limit = request.args.get("limit", "").strip()
    collection_id = request.args.get("collection_id", "").strip()
    difficulty = request.args.get("difficulty", "").strip().lower()
    if not collection_id.isdigit():
        return api_error("collection_id é obrigatório.", 400, code="required_collection")
    if difficulty and difficulty not in ("easy", "hard"):
        return api_error("difficulty inválido.", 400, code="invalid_difficulty")
    if limit and not limit.isdigit():
        return api_error("limit inválido.", 400, code="invalid_limit")
    limit_val = int(limit) if limit.isdigit() else None
    with get_db() as conn:
        if not _collection_exists(conn, int(collection_id)):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        where_clause = ""
        params = []
        where_clause = "WHERE collection_id = ?"
        params.append(int(collection_id))
        if difficulty in ("easy", "hard"):
            where_clause += " AND id IN (SELECT card_id FROM card_difficulty WHERE difficulty = ?)"
            params.append(difficulty)
        if limit_val is not None:
            rows = conn.execute(
                f"""
                SELECT id, collection_id, question, answer, created_at
                FROM cards
                {where_clause}
                ORDER BY RANDOM() LIMIT ?
                """,
                (*params, limit_val),
            ).fetchall()
        else:
            rows = conn.execute(
                f"""
                SELECT id, collection_id, question, answer, created_at
                FROM cards
                {where_clause}
                ORDER BY RANDOM()
                """,
                params,
            ).fetchall()
    return jsonify([row_to_dict(row) for row in rows])


@app.route("/api/study/collections", methods=["GET"])
def list_study_collections():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT
                c.id,
                c.name,
                c.created_at,
                COUNT(cards.id) AS card_count,
                SUM(CASE WHEN cd.difficulty = 'easy' THEN 1 ELSE 0 END) AS easy_count,
                SUM(CASE WHEN cd.difficulty = 'hard' THEN 1 ELSE 0 END) AS hard_count,
                SUM(CASE WHEN cd.difficulty IN ('easy', 'hard') THEN 1 ELSE 0 END) AS rated_count
            FROM collections c
            LEFT JOIN cards ON cards.collection_id = c.id
            LEFT JOIN card_difficulty cd ON cd.card_id = cards.id
            GROUP BY c.id
            ORDER BY c.name ASC
            """
        ).fetchall()
    payload = []
    for row in rows:
        card_count = row["card_count"] or 0
        easy_count = row["easy_count"] or 0
        hard_count = row["hard_count"] or 0
        rated_count = row["rated_count"] or 0
        payload.append(
            {
                "id": row["id"],
                "name": row["name"],
                "created_at": row["created_at"],
                "card_count": card_count,
                "easy_count": easy_count,
                "hard_count": hard_count,
                "difficulty_ready": card_count > 0 and rated_count == card_count,
            }
        )
    return jsonify(payload)


@app.route("/api/study/sessions", methods=["GET"])
def list_study_sessions():
    with get_db() as conn:
        rows = conn.execute(
            """
            SELECT id, created_at, duration_sec, total, correct, incorrect, easy, hard, details_json
            FROM study_sessions
            ORDER BY id DESC
            LIMIT 20
            """
        ).fetchall()
    sessions = []
    for row in rows:
        try:
            details = json.loads(row["details_json"])
        except (TypeError, json.JSONDecodeError):
            details = {}
        sessions.append(
            {
                "id": row["id"],
                "created_at": row["created_at"],
                "duration_sec": row["duration_sec"],
                "total": row["total"],
                "correct": row["correct"],
                "incorrect": row["incorrect"],
                "easy": row["easy"],
                "hard": row["hard"],
                "details": details,
            }
        )
    return jsonify(sessions)


@app.route("/api/study/sessions", methods=["POST"])
def create_study_session():
    data, error = _get_request_json()
    if error:
        return error
    duration_sec, error = _parse_int(data.get("duration_sec"), "duration_sec", min_value=0)
    if error:
        return error
    total, error = _parse_int(data.get("total"), "total", min_value=0)
    if error:
        return error
    correct, error = _parse_int(data.get("correct"), "correct", min_value=0)
    if error:
        return error
    incorrect, error = _parse_int(data.get("incorrect"), "incorrect", min_value=0)
    if error:
        return error
    easy, error = _parse_int(data.get("easy"), "easy", min_value=0)
    if error:
        return error
    hard, error = _parse_int(data.get("hard"), "hard", min_value=0)
    if error:
        return error
    details = data.get("details") if isinstance(data, dict) else {}
    if details is None:
        details = {}
    if not isinstance(details, dict):
        return api_error("details inválido.", 400, code="invalid_details")
    collection_id = data.get("collection_id")
    created_at = datetime.now(timezone.utc).isoformat()
    collection_id_int = int(collection_id) if str(collection_id).isdigit() else None
    if collection_id is not None and collection_id_int is None:
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if total is None:
        total = 0
    if duration_sec is None:
        duration_sec = 0
    if correct is None:
        correct = 0
    if incorrect is None:
        incorrect = 0
    if easy is None:
        easy = 0
    if hard is None:
        hard = 0
    if correct + incorrect > total:
        return api_error("correct + incorrect não pode exceder total.", 400, code="invalid_score")
    if easy + hard > total:
        return api_error("easy + hard não pode exceder total.", 400, code="invalid_score")
    easy_cards = details.get("easy_cards") or []
    hard_cards = details.get("hard_cards") or []
    if not isinstance(easy_cards, list) or not isinstance(hard_cards, list):
        return api_error("details inválido.", 400, code="invalid_details")

    with get_db() as conn:
        if collection_id_int is not None and not _collection_exists(conn, collection_id_int):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        cur = conn.execute(
            """
            INSERT INTO study_sessions
            (collection_id, created_at, duration_sec, total, correct, incorrect, easy, hard, details_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                collection_id_int,
                created_at,
                duration_sec,
                total,
                correct,
                incorrect,
                easy,
                hard,
                json.dumps(details),
            ),
        )
        session_id = cur.lastrowid
        updated_at = created_at
        card_updates = {}
        for card_id in easy_cards:
            if str(card_id).isdigit():
                card_updates[int(card_id)] = "easy"
        for card_id in hard_cards:
            if str(card_id).isdigit():
                card_updates[int(card_id)] = "hard"
        if card_updates:
            conn.executemany(
                """
                INSERT INTO card_difficulty (card_id, difficulty, updated_at)
                VALUES (?, ?, ?)
                ON CONFLICT(card_id) DO UPDATE SET difficulty = excluded.difficulty, updated_at = excluded.updated_at
                """,
                [(card_id, difficulty, updated_at) for card_id, difficulty in card_updates.items()],
            )

    return jsonify(
        {
            "id": session_id,
            "created_at": created_at,
            "duration_sec": duration_sec,
            "total": total,
            "correct": correct,
            "incorrect": incorrect,
            "easy": easy,
            "hard": hard,
        }
    )


@app.route("/api/study/sessions/<int:session_id>", methods=["DELETE"])
def delete_study_session(session_id):
    with get_db() as conn:
        cur = conn.execute("DELETE FROM study_sessions WHERE id = ?", (session_id,))
    if cur.rowcount == 0:
        return api_error("Sessão não encontrada.", 404, code="not_found")
    return jsonify({"ok": True})


@app.route("/api/cards", methods=["POST"])
def create_card():
    data, error = _get_request_json()
    if error:
        return error
    question = (data.get("question") or "").strip()
    answer = (data.get("answer") or "").strip()
    collection_id, error = _parse_int(
        data.get("collection_id"), "collection_id", required=True, min_value=1
    )
    if error:
        return error
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
    if not question or not answer:
        return api_error("Pergunta e resposta são obrigatórias.", 400, code="required_fields")

    created_at = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        cur = conn.execute(
            """
            INSERT INTO cards (collection_id, question, answer, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (collection_id, question, answer, created_at),
        )
        card_id = cur.lastrowid
    return jsonify(
        {
            "id": card_id,
            "collection_id": collection_id,
            "question": question,
            "answer": answer,
            "created_at": created_at,
        }
    )


@app.route("/api/cards/<int:card_id>", methods=["DELETE"])
def delete_card(card_id):
    with get_db() as conn:
        cur = conn.execute("DELETE FROM cards WHERE id = ?", (card_id,))
    if cur.rowcount == 0:
        return api_error("Card não encontrado.", 404, code="not_found")
    return jsonify({"ok": True})


@app.route("/api/generate", methods=["POST"])
def generate_cards():
    api_key = (
        request.headers.get("X-API-Key", "").strip()
        or request.headers.get("X-OpenAI-Key", "").strip()
        or os.environ.get("OPENAI_API_KEY", "").strip()
        or os.environ.get("GEMINI_API_KEY", "").strip()
    )
    if not api_key:
        return api_error(
            "Defina a API Key (OpenAI ou Gemini) nas configurações.", 400, code="missing_api_key"
        )

    data, error = _get_request_json()
    if error:
        return error
    topic = (data.get("topic") or "").strip()
    count_raw = data.get("count")
    if count_raw is None or (isinstance(count_raw, str) and not count_raw.strip()):
        count = 5
    else:
        count, error = _parse_int(count_raw, "count", min_value=1, max_value=60)
        if error:
            return error
    collection_id, error = _parse_int(
        data.get("collection_id"), "collection_id", required=True, min_value=1
    )
    if error:
        return error

    prompt = (
        "Gere flashcards no formato JSON. "
        "Responda SOMENTE com um JSON contendo a chave 'cards' e uma lista de objetos "
        "com 'question' e 'answer'. "
        f"Tema: {topic or 'geral'}. Quantidade: {count}."
    )

    provider = detect_provider(api_key, request.headers.get("X-Provider", ""))
    output_text = ""
    usage = {}
    if provider == "gemini":
        try:
            response = _ai_session.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}",
                headers={"Content-Type": "application/json"},
                json={
                    "contents": [
                        {
                            "role": "user",
                            "parts": [{"text": prompt}],
                        }
                    ],
                    "generationConfig": {"responseMimeType": "application/json"},
                },
                timeout=(AI_CONNECT_TIMEOUT, AI_READ_TIMEOUT),
            )
        except requests.RequestException:
            return api_error(
                "Falha ao conectar com o provedor de IA.", 502, code="ai_network_error"
            )
        if response.status_code in (401, 403):
            return api_error("API Key inválida ou sem permissão.", 401, code="ai_auth_error")
        if response.status_code == 429:
            return api_error("Limite de requisições excedido.", 429, code="ai_rate_limited")
        if response.status_code >= 400:
            return api_error(
                "Falha ao gerar cards.",
                502,
                code="ai_error",
                details=_trim_error_details(response.text),
            )
        response_json = response.json()
        output_text = extract_gemini_text(response_json)
        usage = response_json.get("usageMetadata") or {}
    else:
        try:
            response = _ai_session.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "gpt-4.1",
                    "input": prompt,
                    "text": {
                        "format": {"type": "json_object"},
                    },
                },
                timeout=(AI_CONNECT_TIMEOUT, AI_READ_TIMEOUT),
            )
        except requests.RequestException:
            return api_error(
                "Falha ao conectar com o provedor de IA.", 502, code="ai_network_error"
            )

        if response.status_code in (401, 403):
            return api_error("API Key inválida ou sem permissão.", 401, code="ai_auth_error")
        if response.status_code == 429:
            return api_error("Limite de requisições excedido.", 429, code="ai_rate_limited")
        if response.status_code >= 400:
            return api_error(
                "Falha ao gerar cards.",
                502,
                code="ai_error",
                details=_trim_error_details(response.text),
            )

        response_json = response.json()
        output_text = extract_output_text(response_json)
        usage = response_json.get("usage") or {}

    try:
        payload = json.loads(output_text)
        if not isinstance(payload, dict):
            return api_error("Resposta da IA não veio no formato esperado.", 502, code="ai_invalid_payload")
        cards = payload.get("cards", [])
        if not isinstance(cards, list):
            return api_error("Resposta da IA não veio no formato esperado.", 502, code="ai_invalid_payload")
    except json.JSONDecodeError:
        return api_error(
            "Resposta da IA não veio em JSON válido.",
            502,
            code="ai_invalid_json",
            details=_trim_error_details(output_text),
        )

    saved = []
    created_at = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        for card in cards:
            question = str(card.get("question", "")).strip()
            answer = str(card.get("answer", "")).strip()
            if not question or not answer:
                continue
            cur = conn.execute(
                """
                INSERT INTO cards (collection_id, question, answer, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    int(collection_id),
                    question,
                    answer,
                    created_at,
                ),
            )
            saved.append(
                {
                    "id": cur.lastrowid,
                    "collection_id": collection_id,
                    "question": question,
                    "answer": answer,
                    "created_at": created_at,
                }
            )

    if not saved:
        return api_error("Nenhum card válido foi gerado.", 502, code="ai_empty")
    return jsonify({"created": saved, "usage": usage})


@app.route("/api/exam/generate", methods=["POST"])
def generate_exam():
    api_key = (
        request.headers.get("X-API-Key", "").strip()
        or request.headers.get("X-OpenAI-Key", "").strip()
        or os.environ.get("OPENAI_API_KEY", "").strip()
        or os.environ.get("GEMINI_API_KEY", "").strip()
    )
    if not api_key:
        return api_error(
            "Defina a API Key (OpenAI ou Gemini) nas configurações.", 400, code="missing_api_key"
        )

    data, error = _get_request_json()
    if error:
        return error
    topic = (data.get("topic") or "").strip()
    count_raw = data.get("count")
    if count_raw is None or (isinstance(count_raw, str) and not count_raw.strip()):
        count = 10
    else:
        count, error = _parse_int(count_raw, "count", min_value=1, max_value=40)
        if error:
            return error

    allow_multi_raw = data.get("allow_multi") if isinstance(data, dict) else None
    allow_multi = _coerce_bool(allow_multi_raw)
    allow_multi = bool(allow_multi) if allow_multi is not None else False
    max_correct = 2 if allow_multi else 1
    multi_target = 0
    if allow_multi:
        multi_target = max(1, int(round(count * 0.25))) if count > 1 else 0
        multi_target = min(count, multi_target)

    collection_id = data.get("collection_id")
    collection_id_int = int(collection_id) if str(collection_id).isdigit() else None
    if collection_id is not None and collection_id_int is None:
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if collection_id_int is not None and collection_id_int < 1:
        return api_error("collection_id inválido.", 400, code="invalid_collection")

    reference_block = ""
    collection_name = ""
    if collection_id_int is not None:
        with get_db() as conn:
            if not _collection_exists(conn, collection_id_int):
                return api_error("Coleção não encontrada.", 404, code="collection_not_found")
            collection_row = conn.execute(
                "SELECT name FROM collections WHERE id = ?",
                (collection_id_int,),
            ).fetchone()
            reference_limit = 120
            if count >= 25:
                reference_limit = 160
            rows = conn.execute(
                """
                SELECT question, answer
                FROM cards
                WHERE collection_id = ?
                ORDER BY RANDOM()
                LIMIT ?
                """,
                (collection_id_int, reference_limit),
            ).fetchall()
        collection_name = str(collection_row["name"] or "").strip() if collection_row else ""
        if rows:
            lines = []
            max_chars = 9000
            current_chars = 0
            for idx, row in enumerate(rows, start=1):
                q = str(row["question"] or "").strip()
                a = str(row["answer"] or "").strip()
                if not q or not a:
                    continue
                chunk = f"{idx}. Q: {q}\n   A: {a}"
                if current_chars + len(chunk) > max_chars:
                    break
                lines.append(chunk)
                current_chars += len(chunk)
            if lines:
                reference_block = (
                    f"\nMaterial de estudo (coleção: {collection_name or 'Coleção'}):\n"
                    + "\n".join(lines)
                    + "\n"
                )

    topic_label = topic or collection_name or "geral"

    multi_instruction = ""
    if max_correct == 1:
        multi_instruction = "- 'correct' deve ter EXATAMENTE 1 índice.\n"
    else:
        multi_instruction = (
            f"- Crie aproximadamente {multi_target} perguntas com 2 respostas corretas (ou menos se não fizer sentido).\n"
            "- Para essas, 'correct' deve ter EXATAMENTE 2 índices.\n"
            "- Para as demais, 'correct' deve ter EXATAMENTE 1 índice.\n"
        )

    prompt = (
        "Gere uma prova (questões de múltipla escolha) no formato JSON. "
        "Responda SOMENTE com um JSON contendo a chave 'questions' e uma lista de objetos "
        "com 'question', 'options', 'correct' e 'explanation'.\n"
        f"Tema/assunto: {topic_label}.\n"
        f"Quantidade: {count}.\n"
        f"{reference_block}\n"
        "Formato esperado:\n"
        "{\n"
        '  "questions": [\n'
        "    {\n"
        '      "question": "Pergunta...",\n'
        '      "options": ["Opção 1", "Opção 2", "Opção 3", "Opção 4"],\n'
        '      "correct": [0],\n'
        '      "explanation": "Explicação curta."\n'
        "    }\n"
        "  ]\n"
        "}\n\n"
        "Regras:\n"
        "- Sempre 4 opções em 'options'.\n"
        "- Não coloque letras (A, B, C, D) dentro das opções.\n"
        "- 'correct' deve ser uma lista de índices (0 a 3) que apontam para 'options'.\n"
        "- 'explanation' é obrigatório e deve ser curto (1–2 frases), explicando por que a(s) correta(s) está(ão) correta(s).\n"
        "- A explicação deve citar explicitamente a(s) alternativa(s) correta(s) usando o texto da opção (ex.: 'Amazon S3').\n"
        "- Revise o gabarito: a(s) opção(ões) em 'correct' deve(m) corresponder à explicação.\n"
        f"{multi_instruction}"
        "- Use o material de estudo da coleção como base principal quando ele for fornecido.\n"
        "- Evite alternativas ambíguas e não use 'todas as anteriores'/'nenhuma das anteriores'.\n"
    )

    provider = detect_provider(api_key, request.headers.get("X-Provider", ""))
    output_text = ""
    usage = {}
    if provider == "gemini":
        try:
            response = _ai_session.post(
                f"https://generativelanguage.googleapis.com/v1beta/models/gemini-1.5-flash:generateContent?key={api_key}",
                headers={"Content-Type": "application/json"},
                json={
                    "contents": [
                        {
                            "role": "user",
                            "parts": [{"text": prompt}],
                        }
                    ],
                    "generationConfig": {"responseMimeType": "application/json"},
                },
                timeout=(AI_CONNECT_TIMEOUT, AI_READ_TIMEOUT),
            )
        except requests.RequestException:
            return api_error(
                "Falha ao conectar com o provedor de IA.", 502, code="ai_network_error"
            )
        if response.status_code in (401, 403):
            return api_error("API Key inválida ou sem permissão.", 401, code="ai_auth_error")
        if response.status_code == 429:
            return api_error("Limite de requisições excedido.", 429, code="ai_rate_limited")
        if response.status_code >= 400:
            return api_error(
                "Falha ao gerar prova.",
                502,
                code="ai_error",
                details=_trim_error_details(response.text),
            )
        response_json = response.json()
        output_text = extract_gemini_text(response_json)
        usage = response_json.get("usageMetadata") or {}
    else:
        try:
            response = _ai_session.post(
                "https://api.openai.com/v1/responses",
                headers={
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                },
                json={
                    "model": "gpt-4.1",
                    "input": prompt,
                    "text": {
                        "format": {"type": "json_object"},
                    },
                },
                timeout=(AI_CONNECT_TIMEOUT, AI_READ_TIMEOUT),
            )
        except requests.RequestException:
            return api_error(
                "Falha ao conectar com o provedor de IA.", 502, code="ai_network_error"
            )

        if response.status_code in (401, 403):
            return api_error("API Key inválida ou sem permissão.", 401, code="ai_auth_error")
        if response.status_code == 429:
            return api_error("Limite de requisições excedido.", 429, code="ai_rate_limited")
        if response.status_code >= 400:
            return api_error(
                "Falha ao gerar prova.",
                502,
                code="ai_error",
                details=_trim_error_details(response.text),
            )

        response_json = response.json()
        output_text = extract_output_text(response_json)
        usage = response_json.get("usage") or {}

    try:
        payload = json.loads(output_text)
        if not isinstance(payload, dict):
            return api_error("Resposta da IA não veio no formato esperado.", 502, code="ai_invalid_payload")
        questions = payload.get("questions", [])
        if not isinstance(questions, list):
            return api_error("Resposta da IA não veio no formato esperado.", 502, code="ai_invalid_payload")
    except json.JSONDecodeError:
        return api_error(
            "Resposta da IA não veio em JSON válido.",
            502,
            code="ai_invalid_json",
            details=_trim_error_details(output_text),
        )

    normalized = []
    for item in questions:
        if len(normalized) >= count:
            break
        if not isinstance(item, dict):
            continue
        question_text = str(item.get("question") or "").strip()
        options = item.get("options")
        if not question_text or not isinstance(options, list):
            continue
        options_norm = [str(opt or "").strip() for opt in options]
        options_norm = [opt for opt in options_norm if opt]
        if len(options_norm) != 4:
            continue
        correct_norm = _normalize_mcq_correct(item.get("correct"))
        if not correct_norm:
            continue
        if any(idx < 0 or idx > 3 for idx in correct_norm):
            continue
        if max_correct == 1:
            if len(correct_norm) != 1:
                continue
        else:
            if len(correct_norm) not in (1, 2):
                continue
        explanation = str(item.get("explanation") or "").strip()
        if not explanation:
            continue
        if len(explanation) > 360:
            explanation = explanation[:357] + "..."
        normalized.append(
            {
                "question": question_text,
                "options": options_norm,
                "correct": correct_norm,
                "explanation": explanation,
            }
        )

    normalized = [_fix_mcq_correct_from_explanation(q) for q in normalized]

    if not normalized:
        return api_error("Nenhuma pergunta válida foi gerada.", 502, code="ai_empty")
    return jsonify({"questions": normalized, "usage": usage})


@app.route("/api/exam/sessions", methods=["GET"])
def list_exam_sessions():
    limit, error = _parse_int(request.args.get("limit"), "limit", min_value=1, max_value=200)
    if error:
        return error
    limit_val = limit or 20

    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None

    with get_db() as conn:
        where_clause = ""
        params = []
        if collection_id_val is not None:
            if not _collection_exists(conn, collection_id_val):
                return api_error("Coleção não encontrada.", 404, code="collection_not_found")
            where_clause = "WHERE collection_id = ?"
            params.append(collection_id_val)

        summary = conn.execute(
            f"""
            SELECT
                COUNT(*) AS count,
                COALESCE(SUM(total), 0) AS total,
                COALESCE(SUM(answered), 0) AS answered,
                COALESCE(SUM(correct), 0) AS correct,
                COALESCE(SUM(incorrect), 0) AS incorrect
            FROM exam_sessions
            {where_clause}
            """,
            params,
        ).fetchone()

        rows = conn.execute(
            f"""
            SELECT id, collection_id, topic, name, created_at, started_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json
            FROM exam_sessions
            {where_clause}
            ORDER BY id DESC
            LIMIT ?
            """,
            (*params, limit_val),
        ).fetchall()

    sessions = []
    for row in rows:
        try:
            details = json.loads(row["details_json"])
        except (TypeError, json.JSONDecodeError):
            details = {}
        sessions.append(
            {
                "id": row["id"],
                "collection_id": row["collection_id"],
                "topic": row["topic"] or "",
                "name": row["name"] or "",
                "created_at": row["created_at"],
                "started_at": row["started_at"],
                "duration_sec": row["duration_sec"],
                "time_limit_sec": row["time_limit_sec"],
                "allow_multi": bool(row["allow_multi"]),
                "total": row["total"],
                "answered": row["answered"],
                "correct": row["correct"],
                "incorrect": row["incorrect"],
                "details": details,
            }
        )

    summary_payload = {
        "count": int(summary["count"] or 0) if summary else 0,
        "total": int(summary["total"] or 0) if summary else 0,
        "answered": int(summary["answered"] or 0) if summary else 0,
        "correct": int(summary["correct"] or 0) if summary else 0,
        "incorrect": int(summary["incorrect"] or 0) if summary else 0,
    }
    return jsonify({"sessions": sessions, "summary": summary_payload})


@app.route("/api/exam/sessions", methods=["POST"])
def create_exam_session():
    data, error = _get_request_json()
    if error:
        return error

    duration_sec, error = _parse_int(data.get("duration_sec"), "duration_sec", min_value=0)
    if error:
        return error
    time_limit_sec, error = _parse_int(data.get("time_limit_sec"), "time_limit_sec", min_value=0)
    if error:
        return error
    total, error = _parse_int(data.get("total"), "total", required=True, min_value=1)
    if error:
        return error
    answered, error = _parse_int(data.get("answered"), "answered", required=True, min_value=0)
    if error:
        return error
    correct, error = _parse_int(data.get("correct"), "correct", required=True, min_value=0)
    if error:
        return error
    incorrect, error = _parse_int(data.get("incorrect"), "incorrect", required=True, min_value=0)
    if error:
        return error

    allow_multi = _coerce_bool(data.get("allow_multi"))
    allow_multi_int = 1 if allow_multi else 0

    collection_id = data.get("collection_id")
    collection_id_int = int(collection_id) if str(collection_id).isdigit() else None
    if collection_id is not None and collection_id_int is None:
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if collection_id_int is not None and collection_id_int < 1:
        return api_error("collection_id inválido.", 400, code="invalid_collection")

    topic = str(data.get("topic") or "").strip()
    if len(topic) > 200:
        return api_error("topic inválido.", 400, code="invalid_topic")

    name = str(data.get("name") or "").strip()
    if len(name) > 120:
        return api_error("name inválido.", 400, code="invalid_name")

    if duration_sec is None:
        duration_sec = 0
    if answered > total:
        return api_error("answered não pode exceder total.", 400, code="invalid_score")
    if correct + incorrect != answered:
        return api_error("correct + incorrect deve ser igual a answered.", 400, code="invalid_score")
    if correct + incorrect > total:
        return api_error("correct + incorrect não pode exceder total.", 400, code="invalid_score")

    details = data.get("details") if isinstance(data, dict) else {}
    if details is None:
        details = {}
    if not isinstance(details, dict):
        return api_error("details inválido.", 400, code="invalid_details")

    created_at = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        if collection_id_int is not None and not _collection_exists(conn, collection_id_int):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        cur = conn.execute(
            """
            INSERT INTO exam_sessions
            (collection_id, topic, name, created_at, started_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                collection_id_int,
                topic,
                name,
                created_at,
                created_at,
                duration_sec,
                time_limit_sec,
                allow_multi_int,
                total,
                answered,
                correct,
                incorrect,
                json.dumps(details),
            ),
        )
        session_id = cur.lastrowid
    return jsonify({"id": session_id, "created_at": created_at})


@app.route("/api/exam/sessions/<int:session_id>", methods=["DELETE"])
def delete_exam_session(session_id):
    with get_db() as conn:
        cur = conn.execute("DELETE FROM exam_sessions WHERE id = ?", (session_id,))
    if cur.rowcount == 0:
        return api_error("Prova não encontrada.", 404, code="not_found")
    return jsonify({"ok": True})


@app.route("/api/exam/sessions/<int:session_id>", methods=["PATCH"])
def update_exam_session(session_id):
    data, error = _get_request_json()
    if error:
        return error

    updates = []
    params = []

    has_name = isinstance(data, dict) and "name" in data
    if has_name:
        name = str(data.get("name") or "").strip()
        if len(name) > 120:
            return api_error("name inválido.", 400, code="invalid_name")
        updates.append("name = ?")
        params.append(name)
    else:
        name = None

    mark_started = bool(data.get("mark_started")) if isinstance(data, dict) and "mark_started" in data else False
    started_at = None
    if mark_started:
        started_at = datetime.now(timezone.utc).isoformat()
        updates.append("started_at = COALESCE(started_at, ?)")
        params.append(started_at)

    if not updates:
        return api_error("Nada para atualizar.", 400, code="no_updates")

    with get_db() as conn:
        cur = conn.execute(
            f"UPDATE exam_sessions SET {', '.join(updates)} WHERE id = ?",
            (*params, session_id),
        )
        if cur.rowcount == 0:
            return api_error("Prova não encontrada.", 404, code="not_found")
        row = conn.execute("SELECT name, started_at FROM exam_sessions WHERE id = ?", (session_id,)).fetchone()

    return jsonify({"ok": True, "name": row["name"] if row else "", "started_at": row["started_at"] if row else None})


@app.route("/api/exam/sessions/clear", methods=["POST"])
def clear_exam_sessions():
    data, error = _get_request_json(required=False)
    if error:
        return error
    collection_id = data.get("collection_id") if isinstance(data, dict) else None
    collection_id_int = int(collection_id) if str(collection_id).isdigit() else None
    if collection_id is not None and collection_id_int is None:
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if collection_id_int is not None and collection_id_int < 1:
        return api_error("collection_id inválido.", 400, code="invalid_collection")

    with get_db() as conn:
        if collection_id_int is not None and not _collection_exists(conn, collection_id_int):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        if collection_id_int is not None:
            cur = conn.execute("DELETE FROM exam_sessions WHERE collection_id = ?", (collection_id_int,))
        else:
            cur = conn.execute("DELETE FROM exam_sessions")
        deleted = cur.rowcount or 0
    return jsonify({"ok": True, "deleted": deleted})


def _exam_correct_to_label(indices):
    letters = []
    for idx in indices or []:
        try:
            i = int(idx)
        except (TypeError, ValueError):
            continue
        if 0 <= i <= 3:
            letters.append(("A", "B", "C", "D")[i])
    return ",".join(letters)


def _parse_exam_export_questions(details_json):
    if not details_json:
        return []
    try:
        details = json.loads(details_json)
    except (TypeError, json.JSONDecodeError):
        return []
    if not isinstance(details, dict):
        return []
    questions = details.get("questions") or []
    if not isinstance(questions, list):
        return []
    return [q for q in questions if isinstance(q, dict)]


@app.route("/api/exam/export/csv", methods=["GET"])
def export_exam_csv():
    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None

    with get_db() as conn:
        if collection_id_val is not None and not _collection_exists(conn, collection_id_val):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        if collection_id_val is not None:
            rows = conn.execute(
                """
                SELECT id, collection_id, topic, name, created_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json
                FROM exam_sessions
                WHERE collection_id = ?
                ORDER BY id DESC
                """,
                (collection_id_val,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id, collection_id, topic, name, created_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json
                FROM exam_sessions
                ORDER BY id DESC
                """
            ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(
        [
            "exam_id",
            "name",
            "topic",
            "collection_id",
            "created_at",
            "allow_multi",
            "time_limit_sec",
            "question_index",
            "question",
            "option_1",
            "option_2",
            "option_3",
            "option_4",
            "correct",
            "explanation",
        ]
    )
    for row in rows:
        questions = _parse_exam_export_questions(row["details_json"])
        for q_idx, q in enumerate(questions, start=1):
            question_text = str(q.get("question") or "").strip()
            options = q.get("options") or []
            if not question_text or not isinstance(options, list):
                continue
            options_norm = [str(opt or "").strip() for opt in options]
            if len(options_norm) != 4 or any(not opt for opt in options_norm):
                continue
            correct_norm = _normalize_mcq_correct(q.get("correct"))
            if not correct_norm:
                continue
            explanation = str(q.get("explanation") or "").strip()
            writer.writerow(
                [
                    row["id"],
                    row["name"] or "",
                    row["topic"] or "",
                    row["collection_id"] if row["collection_id"] is not None else "",
                    row["created_at"],
                    int(row["allow_multi"] or 0),
                    row["time_limit_sec"] if row["time_limit_sec"] is not None else "",
                    q_idx,
                    question_text,
                    options_norm[0],
                    options_norm[1],
                    options_norm[2],
                    options_norm[3],
                    _exam_correct_to_label(correct_norm) or "",
                    explanation,
                ]
            )

    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    return send_file(
        mem,
        mimetype="text/csv",
        as_attachment=True,
        download_name="provas.csv",
    )


@app.route("/api/exam/export/xlsx", methods=["GET"])
def export_exam_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        return api_error(
            "Instale openpyxl para exportar XLSX: pip install openpyxl",
            400,
            code="missing_dependency",
        )
    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None

    with get_db() as conn:
        if collection_id_val is not None and not _collection_exists(conn, collection_id_val):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        if collection_id_val is not None:
            rows = conn.execute(
                """
                SELECT id, collection_id, topic, name, created_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json
                FROM exam_sessions
                WHERE collection_id = ?
                ORDER BY id DESC
                """,
                (collection_id_val,),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT id, collection_id, topic, name, created_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json
                FROM exam_sessions
                ORDER BY id DESC
                """
            ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Provas"
    ws.append(
        [
            "exam_id",
            "name",
            "topic",
            "collection_id",
            "created_at",
            "allow_multi",
            "time_limit_sec",
            "question_index",
            "question",
            "option_1",
            "option_2",
            "option_3",
            "option_4",
            "correct",
            "explanation",
        ]
    )
    for row in rows:
        questions = _parse_exam_export_questions(row["details_json"])
        for q_idx, q in enumerate(questions, start=1):
            question_text = str(q.get("question") or "").strip()
            options = q.get("options") or []
            if not question_text or not isinstance(options, list):
                continue
            options_norm = [str(opt or "").strip() for opt in options]
            if len(options_norm) != 4 or any(not opt for opt in options_norm):
                continue
            correct_norm = _normalize_mcq_correct(q.get("correct"))
            if not correct_norm:
                continue
            explanation = str(q.get("explanation") or "").strip()
            ws.append(
                [
                    row["id"],
                    row["name"] or "",
                    row["topic"] or "",
                    row["collection_id"] if row["collection_id"] is not None else "",
                    row["created_at"],
                    int(row["allow_multi"] or 0),
                    row["time_limit_sec"] if row["time_limit_sec"] is not None else "",
                    q_idx,
                    question_text,
                    options_norm[0],
                    options_norm[1],
                    options_norm[2],
                    options_norm[3],
                    _exam_correct_to_label(correct_norm) or "",
                    explanation,
                ]
            )

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="provas.xlsx",
    )


def _normalize_exam_import_header(value):
    return str(value or "").strip().lower().replace(" ", "_")


def _parse_exam_import_rows_csv(file_storage):
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")
    reader = csv.DictReader(io.StringIO(text))
    rows = []
    for row in reader:
        if not isinstance(row, dict):
            continue
        cleaned = {}
        for key, val in row.items():
            cleaned[_normalize_exam_import_header(key)] = val
        rows.append(cleaned)
    return rows


def _parse_exam_import_rows_xlsx(file_storage):
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("missing_dependency") from exc
    stream = getattr(file_storage, "stream", file_storage)
    try:
        stream.seek(0)
    except Exception:
        pass
    wb = load_workbook(stream, read_only=True, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    header = [_normalize_exam_import_header(h) for h in list(all_rows[0])]
    rows = []
    for raw_row in all_rows[1:]:
        if not raw_row:
            continue
        data = {}
        for idx, key in enumerate(header):
            if not key:
                continue
            if idx >= len(raw_row):
                continue
            data[key] = raw_row[idx]
        rows.append(data)
    return rows


def _parse_exam_import_rows_json(file_storage):
    raw = file_storage.read()
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        text = raw.decode("utf-8", errors="replace")
    try:
        payload = json.loads(text)
    except json.JSONDecodeError:
        return []

    def to_row(question, meta, q_idx, exam_id_value):
        if not isinstance(question, dict):
            return None
        q_text = str(question.get("question") or question.get("pergunta") or "").strip()
        if not q_text:
            return None
        options = question.get("options")
        if isinstance(options, list) and len(options) == 4:
            option_1, option_2, option_3, option_4 = [str(o or "").strip() for o in options]
        else:
            option_1 = str(question.get("option_1") or question.get("a") or question.get("alternativa_a") or "").strip()
            option_2 = str(question.get("option_2") or question.get("b") or question.get("alternativa_b") or "").strip()
            option_3 = str(question.get("option_3") or question.get("c") or question.get("alternativa_c") or "").strip()
            option_4 = str(question.get("option_4") or question.get("d") or question.get("alternativa_d") or "").strip()
        if not option_1 or not option_2 or not option_3 or not option_4:
            return None
        correct = question.get("correct")
        explanation = str(question.get("explanation") or question.get("explicacao") or "").strip()
        row = {
            "exam_id": exam_id_value,
            "name": meta.get("name") or meta.get("nome") or "",
            "topic": meta.get("topic") or meta.get("tema") or "",
            "collection_id": meta.get("collection_id") or meta.get("colecao_id") or meta.get("collection") or "",
            "created_at": meta.get("created_at") or meta.get("criado_em") or "",
            "allow_multi": meta.get("allow_multi") if "allow_multi" in meta else meta.get("multi") or meta.get("permitir_duas") or "",
            "time_limit_sec": meta.get("time_limit_sec") or meta.get("tempo") or meta.get("time_limit") or "",
            "question_index": q_idx,
            "question": q_text,
            "option_1": option_1,
            "option_2": option_2,
            "option_3": option_3,
            "option_4": option_4,
            "correct": correct,
            "explanation": explanation,
        }
        return row

    def rows_from_exam(exam_obj, fallback_exam_id):
        if not isinstance(exam_obj, dict):
            return []
        meta = exam_obj
        exam_id_value = (
            str(meta.get("exam_id") or meta.get("id") or meta.get("prova_id") or fallback_exam_id or "").strip()
        )
        questions = exam_obj.get("questions") or exam_obj.get("perguntas") or []
        if not isinstance(questions, list):
            return []
        out = []
        for q_idx, q in enumerate(questions, start=1):
            row = to_row(q, meta, q_idx, exam_id_value)
            if row:
                out.append(row)
        return out

    if isinstance(payload, list):
        # Array of questions
        exam_meta = {}
        out = []
        for q_idx, q in enumerate(payload, start=1):
            row = to_row(q, exam_meta, q_idx, "0")
            if row:
                out.append(row)
        return out

    if isinstance(payload, dict):
        if isinstance(payload.get("exams"), list):
            out = []
            for idx, exam_obj in enumerate(payload["exams"], start=1):
                out.extend(rows_from_exam(exam_obj, str(idx)))
            return out
        # Single exam object with questions
        if isinstance(payload.get("questions"), list) or isinstance(payload.get("perguntas"), list):
            return rows_from_exam(payload, "0")

    return []


def _coerce_exam_id(raw):
    value = str(raw or "").strip()
    return value or None


def _parse_exam_import_int(raw):
    if raw is None or (isinstance(raw, str) and not raw.strip()):
        return None
    try:
        return int(raw)
    except (TypeError, ValueError):
        return None


@app.route("/api/exam/import", methods=["POST"])
def import_exams():
    if "file" not in request.files:
        return api_error("Arquivo é obrigatório.", 400, code="required_file")
    file_storage = request.files["file"]
    filename = (file_storage.filename or "").lower()
    ext = os.path.splitext(filename)[1]

    collection_override_raw = request.form.get("collection_id")
    collection_override = _parse_exam_import_int(collection_override_raw)
    if collection_override_raw is not None and collection_override is None:
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if collection_override is not None and collection_override < 1:
        return api_error("collection_id inválido.", 400, code="invalid_collection")

    try:
        if ext == ".json":
            rows = _parse_exam_import_rows_json(file_storage)
        elif ext == ".xlsx":
            rows = _parse_exam_import_rows_xlsx(file_storage)
        else:
            rows = _parse_exam_import_rows_csv(file_storage)
    except RuntimeError as err:
        if str(err) == "missing_dependency":
            return api_error(
                "Instale openpyxl para importar XLSX: pip install openpyxl",
                400,
                code="missing_dependency",
            )
        return api_error("Falha ao ler arquivo.", 400, code="invalid_file")

    if not rows:
        return api_error("Arquivo vazio ou inválido.", 400, code="invalid_file")

    groups = {}
    for row in rows:
        exam_id = _coerce_exam_id(row.get("exam_id") or row.get("id") or row.get("prova_id"))
        groups.setdefault(exam_id or "0", []).append(row)

    imported = 0
    skipped = 0
    now = datetime.now(timezone.utc).isoformat()
    max_exams = 300
    max_questions = 8000

    with get_db() as conn:
        if collection_override is not None and not _collection_exists(conn, collection_override):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")

        for _, group_rows in list(groups.items())[:max_exams]:
            group_rows_sorted = sorted(
                group_rows,
                key=lambda r: _parse_exam_import_int(r.get("question_index") or r.get("q_index")) or 10**9,
            )
            meta = group_rows_sorted[0] if group_rows_sorted else {}
            topic = str(meta.get("topic") or "").strip()
            name = str(meta.get("name") or "").strip()
            created_at = str(meta.get("created_at") or "").strip() or now
            try:
                if created_at:
                    datetime.fromisoformat(created_at.replace("Z", "+00:00"))
            except ValueError:
                created_at = now

            collection_id = _parse_exam_import_int(meta.get("collection_id"))
            if collection_override is not None:
                collection_id = collection_override
            if collection_id is not None and not _collection_exists(conn, collection_id):
                collection_id = None

            allow_multi = _coerce_bool(meta.get("allow_multi"))
            time_limit_sec = _parse_exam_import_int(meta.get("time_limit_sec"))
            if time_limit_sec is not None and time_limit_sec < 0:
                time_limit_sec = None

            questions = []
            for row in group_rows_sorted:
                if len(questions) >= max_questions:
                    break
                question_text = str(row.get("question") or row.get("pergunta") or "").strip()
                if not question_text:
                    continue
                options = []
                for idx in range(1, 5):
                    opt = (
                        row.get(f"option_{idx}")
                        or row.get(f"option{idx}")
                        or row.get(("a", "b", "c", "d")[idx - 1])
                        or row.get(f"alternativa_{('a','b','c','d')[idx-1]}")
                        or row.get(f"alternativa{idx}")
                        or row.get(f"opcao_{idx}")
                        or row.get(f"opcao{idx}")
                        or row.get(f"opção_{idx}")
                        or row.get(f"opção{idx}")
                    )
                    options.append(str(opt or "").strip())
                if any(not opt for opt in options):
                    continue
                correct_norm = _normalize_mcq_correct(
                    row.get("correct") or row.get("gabarito") or row.get("resposta") or row.get("answer")
                )
                if not correct_norm:
                    continue
                explanation = str(row.get("explanation") or row.get("explicacao") or row.get("comentario") or "").strip()
                questions.append(
                    {
                        "question": question_text,
                        "options": options,
                        "correct": correct_norm,
                        "explanation": explanation,
                    }
                )

            if not questions:
                skipped += 1
                continue

            if allow_multi is None:
                allow_multi = any(len(q.get("correct") or []) > 1 for q in questions)

            total = len(questions)
            details = {"questions": questions, "answers": [[] for _ in range(total)], "usage": {}}
            details["imported"] = True

            conn.execute(
                """
                INSERT INTO exam_sessions
                (collection_id, topic, name, created_at, started_at, duration_sec, time_limit_sec, allow_multi, total, answered, correct, incorrect, details_json)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    collection_id,
                    topic,
                    name,
                    created_at,
                    None,
                    0,
                    time_limit_sec,
                    1 if allow_multi else 0,
                    total,
                    0,
                    0,
                    0,
                    json.dumps(details),
                ),
            )
            imported += 1

    return jsonify({"imported": imported, "skipped": skipped})


@app.route("/api/export/csv", methods=["GET"])
def export_csv():
    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None
    with get_db() as conn:
        if collection_id_val is not None:
            if not _collection_exists(conn, collection_id_val):
                return api_error("Coleção não encontrada.", 404, code="collection_not_found")
            rows = conn.execute(
                """
                SELECT id, collection_id, question, answer, created_at
                FROM cards
                WHERE collection_id = ?
                ORDER BY id DESC
                """,
                (collection_id_val,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, collection_id, question, answer, created_at FROM cards ORDER BY id DESC"
            ).fetchall()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(["id", "collection_id", "question", "answer", "created_at"])
    for row in rows:
        writer.writerow(
            [row["id"], row["collection_id"], row["question"], row["answer"], row["created_at"]]
        )

    mem = io.BytesIO(output.getvalue().encode("utf-8"))
    return send_file(
        mem,
        mimetype="text/csv",
        as_attachment=True,
        download_name="flashcards.csv",
    )


@app.route("/api/export/xlsx", methods=["GET"])
def export_xlsx():
    try:
        from openpyxl import Workbook
    except ImportError:
        return api_error(
            "Instale openpyxl para exportar XLSX: pip install openpyxl",
            400,
            code="missing_dependency",
        )
    collection_id = request.args.get("collection_id", "").strip()
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    collection_id_val = int(collection_id) if collection_id.isdigit() else None
    with get_db() as conn:
        if collection_id_val is not None:
            if not _collection_exists(conn, collection_id_val):
                return api_error("Coleção não encontrada.", 404, code="collection_not_found")
            rows = conn.execute(
                """
                SELECT id, collection_id, question, answer, created_at
                FROM cards
                WHERE collection_id = ?
                ORDER BY id DESC
                """,
                (collection_id_val,),
            ).fetchall()
        else:
            rows = conn.execute(
                "SELECT id, collection_id, question, answer, created_at FROM cards ORDER BY id DESC"
            ).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Flashcards"
    ws.append(["id", "collection_id", "question", "answer", "created_at"])
    for row in rows:
        ws.append([row["id"], row["collection_id"], row["question"], row["answer"], row["created_at"]])

    mem = io.BytesIO()
    wb.save(mem)
    mem.seek(0)
    return send_file(
        mem,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="flashcards.xlsx",
    )


@app.route("/api/collections", methods=["GET"])
def list_collections():
    with get_db() as conn:
        rows = conn.execute(
            """
            WITH card_counts AS (
                SELECT collection_id, COUNT(*) AS card_count
                FROM cards
                GROUP BY collection_id
            ),
            completion_counts AS (
                SELECT ss.collection_id, COUNT(*) AS completion_count
                FROM study_sessions ss
                JOIN card_counts cc ON cc.collection_id = ss.collection_id
                WHERE ss.total >= cc.card_count AND cc.card_count > 0
                GROUP BY ss.collection_id
            )
            SELECT
                c.id,
                c.name,
                c.created_at,
                COALESCE(cc.card_count, 0) AS card_count,
                COALESCE(comp.completion_count, 0) AS completion_count
            FROM collections c
            LEFT JOIN card_counts cc ON cc.collection_id = c.id
            LEFT JOIN completion_counts comp ON comp.collection_id = c.id
            ORDER BY c.name ASC
            """
        ).fetchall()
    return jsonify(
        [
            {
                "id": row["id"],
                "name": row["name"],
                "created_at": row["created_at"],
                "card_count": row["card_count"],
                "completion_count": row["completion_count"],
            }
            for row in rows
        ]
    )


@app.route("/api/collections", methods=["POST"])
def create_collection():
    data, error = _get_request_json()
    if error:
        return error
    name = (data.get("name") or "").strip()
    if not name:
        return api_error("Nome da coleção é obrigatório.", 400, code="required_fields")
    if len(name) > 80:
        return api_error("Nome da coleção é muito longo.", 400, code="invalid_name")
    created_at = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        try:
            cur = conn.execute(
                "INSERT INTO collections (name, created_at) VALUES (?, ?)",
                (name, created_at),
            )
        except sqlite3.IntegrityError:
            return api_error("Já existe uma coleção com esse nome.", 400, code="duplicate_name")
    return jsonify({"id": cur.lastrowid, "name": name, "created_at": created_at})


@app.route("/api/collections/<int:collection_id>", methods=["DELETE"])
def delete_collection(collection_id):
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        conn.execute("DELETE FROM cards WHERE collection_id = ?", (collection_id,))
        conn.execute("DELETE FROM study_goals WHERE collection_id = ?", (collection_id,))
        conn.execute("DELETE FROM collections WHERE id = ?", (collection_id,))
    return jsonify({"ok": True})


def _collection_exists(conn, collection_id):
    row = conn.execute(
        "SELECT id FROM collections WHERE id = ?",
        (collection_id,),
    ).fetchone()
    return row is not None


def _generate_share_token():
    return secrets.token_urlsafe(8)


def _generate_manage_token():
    return secrets.token_urlsafe(16)


def _hash_password(password):
    if not password:
        return ""
    salt = secrets.token_hex(8)
    digest = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return f"{salt}${digest}"


def _verify_password(stored_hash, password):
    if not stored_hash:
        return True
    if not password:
        return False
    try:
        salt, digest = stored_hash.split("$", 1)
    except ValueError:
        return False
    check = hashlib.sha256((salt + password).encode("utf-8")).hexdigest()
    return secrets.compare_digest(check, digest)


def _parse_iso_datetime(raw):
    if not raw:
        return None
    try:
        dt = datetime.fromisoformat(raw)
    except ValueError:
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(timezone.utc)


def _share_is_expired(row):
    expires_at = _parse_iso_datetime(row["expires_at"])
    if not expires_at:
        return False
    return datetime.now(timezone.utc) >= expires_at


def _share_uses_remaining(row):
    max_uses = row["max_uses"]
    if max_uses is None:
        return None
    used = row["uses"] or 0
    return max(0, int(max_uses) - int(used))


def _build_share_payload(conn, collection_id):
    collection = conn.execute(
        "SELECT id, name, created_at FROM collections WHERE id = ?",
        (collection_id,),
    ).fetchone()
    cards = conn.execute(
        """
        SELECT question, answer
        FROM cards
        WHERE collection_id = ?
        ORDER BY id ASC
        """,
        (collection_id,),
    ).fetchall()
    return {
        "collection": {
            "id": collection["id"],
            "name": collection["name"],
            "created_at": collection["created_at"],
        },
        "cards": [{"question": row["question"], "answer": row["answer"]} for row in cards],
    }


def _extract_collection_id_from_payload(payload_json):
    if not payload_json:
        return None
    try:
        payload = json.loads(payload_json)
    except (TypeError, json.JSONDecodeError):
        return None
    if not isinstance(payload, dict):
        return None
    collection = payload.get("collection") or {}
    collection_id = collection.get("id")
    if str(collection_id).isdigit():
        return int(collection_id)
    return None


def _get_focus_summary(conn, collection_id=None):
    rows = []
    params = []
    query = "SELECT created_at, total, correct, incorrect FROM study_sessions"
    if collection_id is not None:
        query += " WHERE collection_id = ?"
        params.append(collection_id)
    rows.extend(conn.execute(query, params).fetchall())

    params = []
    query = "SELECT created_at, total, correct, incorrect FROM exam_sessions"
    if collection_id is not None:
        query += " WHERE collection_id = ?"
        params.append(collection_id)
    rows.extend(conn.execute(query, params).fetchall())
    now = datetime.now(timezone.utc)
    cutoff = now - timedelta(days=7)
    ratios = []
    sessions_count = 0
    for row in rows:
        raw = row["created_at"]
        if not raw:
            continue
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        if dt < cutoff:
            continue
        total = int(row["total"] or 0)
        if total <= 0:
            continue
        answered = int(row["correct"] or 0) + int(row["incorrect"] or 0)
        ratio = min(1.0, answered / total) if total else 0
        ratios.append(ratio)
        sessions_count += 1
    score = round((sum(ratios) / len(ratios)) * 100) if ratios else 0
    if sessions_count == 0:
        label = "Sem sessões recentes"
    elif score >= 85:
        label = "Foco alto"
    elif score >= 60:
        label = "Foco médio"
    else:
        label = "Foco baixo"
    return {
        "score": score,
        "label": label,
        "sessions": sessions_count,
        "period_start": cutoff.date().isoformat(),
        "period_end": now.date().isoformat(),
    }


@app.route("/api/collections/<int:collection_id>/logs", methods=["GET"])
def collection_logs(collection_id):
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        study_row = conn.execute(
            """
            SELECT
                SUM(CASE WHEN total > 0 AND (correct + incorrect) = total THEN 1 ELSE 0 END) AS sessions_complete,
                SUM(CASE WHEN total > 0 AND (correct + incorrect) = total THEN total ELSE 0 END) AS cards_solved,
                SUM(CASE WHEN total > 0 AND (correct + incorrect) = total THEN duration_sec ELSE 0 END) AS total_seconds
            FROM study_sessions
            WHERE collection_id = ?
            """,
            (collection_id,),
        ).fetchone()
        exam_row = conn.execute(
            """
            SELECT
                COUNT(*) AS exams_count,
                COALESCE(SUM(answered), 0) AS answered,
                COALESCE(SUM(correct), 0) AS correct,
                COALESCE(SUM(incorrect), 0) AS incorrect,
                COALESCE(SUM(duration_sec), 0) AS total_seconds
            FROM exam_sessions
            WHERE collection_id = ?
            """,
            (collection_id,),
        ).fetchone()
    return jsonify(
        {
            "collection_id": collection_id,
            "sessions_complete": study_row["sessions_complete"] or 0,
            "cards_solved": study_row["cards_solved"] or 0,
            "total_seconds": study_row["total_seconds"] or 0,
            "exams_count": int(exam_row["exams_count"] or 0) if exam_row else 0,
            "exam_answered": int(exam_row["answered"] or 0) if exam_row else 0,
            "exam_correct": int(exam_row["correct"] or 0) if exam_row else 0,
            "exam_incorrect": int(exam_row["incorrect"] or 0) if exam_row else 0,
            "exam_total_seconds": int(exam_row["total_seconds"] or 0) if exam_row else 0,
        }
    )


@app.route("/api/share", methods=["POST"])
def create_share_link():
    data, error = _get_request_json()
    if error:
        return error
    collection_id, error = _parse_int(
        data.get("collection_id"), "collection_id", required=True, min_value=1
    )
    if error:
        return error
    share_type = (data.get("type") or "").strip().lower()
    password = (data.get("password") or "").strip()
    max_uses_raw = data.get("max_uses")
    if share_type not in ("template", "challenge"):
        return api_error("Tipo inválido.", 400, code="invalid_type")
    created_at = datetime.now(timezone.utc).isoformat()
    expires_at = None
    if share_type == "challenge":
        expires_at = (datetime.now(timezone.utc) + timedelta(minutes=20)).isoformat()
    if max_uses_raw is None or (isinstance(max_uses_raw, str) and not max_uses_raw.strip()):
        max_uses_val = None
    else:
        max_uses_val, error = _parse_int(max_uses_raw, "max_uses", min_value=1)
        if error:
            return error
    if share_type == "challenge" and max_uses_val is None:
        max_uses_val = 1
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        payload = _build_share_payload(conn, collection_id)
        payload["type"] = share_type
        payload["created_at"] = created_at
        token = _generate_share_token()
        manage_token = _generate_manage_token()
        password_hash = _hash_password(password)
        conn.execute(
            """
            INSERT INTO share_links
            (token, type, name, payload_json, created_at, expires_at, max_uses, uses, password_hash, disabled, manage_token, session_token, collection_id)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (
                token,
                share_type,
                payload["collection"]["name"],
                json.dumps(payload),
                created_at,
                expires_at,
                max_uses_val,
                0,
                password_hash,
                0,
                manage_token,
                None,
                collection_id,
            ),
        )
    base_url = request.host_url.rstrip("/")
    url_path = "share" if share_type == "template" else "challenge"
    return jsonify(
        {
            "token": token,
            "url": f"{base_url}/{url_path}/{token}",
            "manage_token": manage_token,
            "expires_at": expires_at,
            "max_uses": max_uses_val,
        }
    )


@app.route("/api/share/list", methods=["GET"])
def list_share_links():
    collection_id = request.args.get("collection_id", "").strip()
    if not collection_id.isdigit():
        return api_error("collection_id é obrigatório.", 400, code="required_collection")
    collection_id_val = int(collection_id)
    with get_db() as conn:
        if not _collection_exists(conn, collection_id_val):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        rows = conn.execute(
            """
            SELECT token, type, name, payload_json, created_at, expires_at, max_uses, uses, password_hash, disabled, collection_id
            FROM share_links
            ORDER BY created_at DESC
            """
        ).fetchall()
        payload = []
        for row in rows:
            row_collection_id = row["collection_id"]
            if row_collection_id is None:
                inferred_id = _extract_collection_id_from_payload(row["payload_json"])
                if inferred_id is not None:
                    row_collection_id = inferred_id
                    conn.execute(
                        "UPDATE share_links SET collection_id = ? WHERE token = ?",
                        (row_collection_id, row["token"]),
                    )
            if row_collection_id != collection_id_val:
                continue
            remaining = _share_uses_remaining(row)
            expired = _share_is_expired(row)
            status = "active"
            if row["disabled"]:
                status = "disabled"
            elif expired:
                status = "expired"
            elif remaining is not None and remaining <= 0:
                status = "exhausted"
            payload.append(
                {
                    "token": row["token"],
                    "type": row["type"],
                    "name": row["name"],
                    "created_at": row["created_at"],
                    "expires_at": row["expires_at"],
                    "max_uses": row["max_uses"],
                    "uses": row["uses"] or 0,
                    "uses_remaining": remaining,
                    "requires_password": bool(row["password_hash"]),
                    "disabled": bool(row["disabled"]),
                    "expired": expired,
                    "status": status,
                }
            )
    return jsonify(payload)


@app.route("/api/share/<string:token>", methods=["GET"])
def get_share_link(token):
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, type, name, payload_json, created_at, expires_at, max_uses, uses, password_hash, disabled
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
    if not row:
        return api_error("Link não encontrado.", 404, code="not_found")
    if row["disabled"]:
        return api_error("Link desativado.", 410, code="disabled")
    if _share_is_expired(row):
        return api_error("Link expirado.", 410, code="expired")
    remaining = _share_uses_remaining(row)
    if remaining is not None and remaining <= 0:
        return api_error("Limite de acessos atingido.", 410, code="exhausted")
    return jsonify(
        {
            "token": row["token"],
            "type": row["type"],
            "name": row["name"],
            "created_at": row["created_at"],
            "expires_at": row["expires_at"],
            "max_uses": row["max_uses"],
            "uses": row["uses"] or 0,
            "uses_remaining": remaining,
            "requires_password": bool(row["password_hash"]),
        }
    )


@app.route("/api/share/<string:token>/access", methods=["POST"])
def access_share_link(token):
    data, error = _get_request_json(required=False)
    if error:
        return error
    password = (data.get("password") or "").strip()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, type, name, payload_json, created_at, expires_at, max_uses, uses, password_hash, disabled, session_token
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return api_error("Link não encontrado.", 404, code="not_found")
        if row["disabled"]:
            return api_error("Link desativado.", 410, code="disabled")
        if _share_is_expired(row):
            return api_error("Link expirado.", 410, code="expired")
        remaining = _share_uses_remaining(row)
        if remaining is not None and remaining <= 0:
            return api_error("Limite de acessos atingido.", 410, code="exhausted")
        if not _verify_password(row["password_hash"], password):
            return api_error("Senha inválida.", 401, code="invalid_password")
        if row["type"] == "challenge" and row["session_token"]:
            return api_error("Desafio já iniciado.", 409, code="already_started")
        try:
            payload = json.loads(row["payload_json"])
        except json.JSONDecodeError:
            return api_error("Payload inválido.", 500, code="invalid_payload")
        new_uses = (row["uses"] or 0) + 1
        session_token = row["session_token"]
        if row["type"] == "challenge":
            session_token = secrets.token_urlsafe(12)
        conn.execute(
            """
            UPDATE share_links
            SET uses = ?, session_token = ?
            WHERE token = ?
            """,
            (new_uses, session_token, token),
        )
    payload["session_token"] = session_token
    return jsonify(payload)


@app.route("/api/share/<string:token>/complete", methods=["POST"])
def complete_share_challenge(token):
    data, error = _get_request_json(required=True)
    if error:
        return error
    session_token = (data.get("session_token") or "").strip()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, type, disabled, session_token
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return api_error("Link não encontrado.", 404, code="not_found")
        if row["type"] != "challenge":
            return api_error("Este link não é um desafio.", 400, code="invalid_type")
        if row["disabled"]:
            return api_error("Link desativado.", 410, code="disabled")
        if not session_token or session_token != row["session_token"]:
            return api_error("Sessão inválida.", 403, code="invalid_session")
        conn.execute(
            "UPDATE share_links SET disabled = 1 WHERE token = ?",
            (token,),
        )
    return jsonify({"ok": True})


@app.route("/api/share/<string:token>/disable", methods=["POST"])
def disable_share_link(token):
    data, error = _get_request_json(required=True)
    if error:
        return error
    manage_token = (data.get("manage_token") or "").strip()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, manage_token
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return api_error("Link não encontrado.", 404, code="not_found")
        if not manage_token or manage_token != (row["manage_token"] or ""):
            return api_error("Permissão negada.", 403, code="forbidden")
        conn.execute("UPDATE share_links SET disabled = 1 WHERE token = ?", (token,))
    return jsonify({"ok": True})


@app.route("/api/share/<string:token>/delete", methods=["POST"])
def delete_share_link(token):
    data, error = _get_request_json(required=True)
    if error:
        return error
    manage_token = (data.get("manage_token") or "").strip()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, manage_token
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return api_error("Link não encontrado.", 404, code="not_found")
        if not manage_token or manage_token != (row["manage_token"] or ""):
            return api_error("Permissão negada.", 403, code="forbidden")
        conn.execute("DELETE FROM share_links WHERE token = ?", (token,))
    return jsonify({"ok": True})


@app.route("/api/share/<string:token>/expire", methods=["POST"])
def expire_share_link(token):
    data, error = _get_request_json()
    if error:
        return error
    manage_token = (data.get("manage_token") or "").strip()
    expires_at_raw = (data.get("expires_at") or "").strip()
    if expires_at_raw:
        expires_at_dt = _parse_iso_datetime(expires_at_raw)
        if not expires_at_dt:
            return api_error("expires_at inválido.", 400, code="invalid_expires_at")
        expires_at = expires_at_dt.isoformat()
    else:
        expires_at = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        row = conn.execute(
            """
            SELECT token, manage_token, disabled
            FROM share_links
            WHERE token = ?
            """,
            (token,),
        ).fetchone()
        if not row:
            return api_error("Link não encontrado.", 404, code="not_found")
        if not manage_token or manage_token != (row["manage_token"] or ""):
            return api_error("Permissão negada.", 403, code="forbidden")
        if row["disabled"]:
            return api_error("Link desativado.", 410, code="disabled")
        conn.execute(
            "UPDATE share_links SET expires_at = ? WHERE token = ?",
            (expires_at, token),
        )
    return jsonify({"ok": True, "expires_at": expires_at})


@app.route("/api/study/summary", methods=["GET"])
def study_summary():
    collection_id = request.args.get("collection_id", "").strip()
    collection_id_val = None
    if collection_id and not collection_id.isdigit():
        return api_error("collection_id inválido.", 400, code="invalid_collection")
    if collection_id.isdigit():
        collection_id_val = int(collection_id)
    with get_db() as conn:
        if collection_id_val is not None and not _collection_exists(conn, collection_id_val):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        summary = _get_focus_summary(conn, collection_id_val)
    summary["collection_id"] = collection_id_val
    return jsonify(summary)


def _normalize_goal_days(days):
    if not isinstance(days, list):
        return []
    cleaned = []
    for day in days:
        try:
            value = int(day)
        except (TypeError, ValueError):
            continue
        if 0 <= value <= 6:
            cleaned.append(value)
    return sorted(set(cleaned))


def _normalize_goal_targets(targets):
    if not isinstance(targets, dict):
        targets = {}

    def to_int(value):
        try:
            val = int(value)
        except (TypeError, ValueError):
            return 0
        return max(0, val)

    return {
        "weekly_cards": to_int(targets.get("weekly_cards")),
        "weekly_minutes": to_int(targets.get("weekly_minutes")),
        "weekly_sessions": to_int(targets.get("weekly_sessions")),
    }


def _parse_goals_json(raw_json):
    try:
        payload = json.loads(raw_json)
    except (TypeError, json.JSONDecodeError):
        return [], _normalize_goal_targets({})
    if isinstance(payload, dict):
        days = payload.get("days") or []
        targets = payload.get("targets") or {}
    else:
        days = payload
        targets = {}
    return _normalize_goal_days(days), _normalize_goal_targets(targets)


def _get_goal_progress(conn, collection_id):
    rows = conn.execute(
        """
        SELECT created_at, duration_sec, total
        FROM study_sessions
        WHERE collection_id = ?
        """,
        (collection_id,),
    ).fetchall()
    today = datetime.now(timezone.utc).date()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    week_sessions = 0
    week_cards = 0
    week_seconds = 0
    session_dates = set()
    for row in rows:
        raw = row["created_at"]
        if not raw:
            continue
        try:
            dt = datetime.fromisoformat(raw)
        except ValueError:
            continue
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        else:
            dt = dt.astimezone(timezone.utc)
        session_date = dt.date()
        session_dates.add(session_date)
        if week_start <= session_date <= week_end:
            week_sessions += 1
            week_cards += int(row["total"] or 0)
            week_seconds += int(row["duration_sec"] or 0)
    week_minutes = int(math.ceil(week_seconds / 60)) if week_seconds else 0
    streak = 0
    cursor = today
    while cursor in session_dates:
        streak += 1
        cursor = cursor - timedelta(days=1)
    return (
        {
            "week_start": week_start.isoformat(),
            "week_end": week_end.isoformat(),
            "sessions": week_sessions,
            "cards": week_cards,
            "minutes": week_minutes,
        },
        streak,
    )


@app.route("/api/goals/<int:collection_id>", methods=["GET"])
def get_goals(collection_id):
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        row = conn.execute(
            """
            SELECT collection_id, days_json, created_at, updated_at
            FROM study_goals
            WHERE collection_id = ?
            """,
            (collection_id,),
        ).fetchone()
        progress, streak = _get_goal_progress(conn, collection_id)
    if not row:
        return jsonify(
            {
                "collection_id": collection_id,
                "days": [],
                "targets": _normalize_goal_targets({}),
                "progress": progress,
                "streak_days": streak,
            }
        )
    days, targets = _parse_goals_json(row["days_json"])
    return jsonify(
        {
            "collection_id": row["collection_id"],
            "days": days,
            "targets": targets,
            "progress": progress,
            "streak_days": streak,
            "created_at": row["created_at"],
            "updated_at": row["updated_at"],
        }
    )


@app.route("/api/goals/<int:collection_id>", methods=["PUT"])
def save_goals(collection_id):
    data, error = _get_request_json()
    if error:
        return error
    days = data.get("days") if isinstance(data, dict) else []
    if days is None:
        days = []
    if not isinstance(days, list):
        return api_error("Formato inválido para dias.", 400, code="invalid_days")
    days = _normalize_goal_days(days)
    targets = _normalize_goal_targets(data.get("targets") if isinstance(data, dict) else {})
    now = datetime.now(timezone.utc).isoformat()
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        existing = conn.execute(
            "SELECT id FROM study_goals WHERE collection_id = ?",
            (collection_id,),
        ).fetchone()
        if existing:
            conn.execute(
                """
                UPDATE study_goals
                SET days_json = ?, updated_at = ?
                WHERE collection_id = ?
                """,
                (json.dumps({"days": days, "targets": targets}), now, collection_id),
            )
        else:
            conn.execute(
                """
                INSERT INTO study_goals (collection_id, days_json, created_at, updated_at)
                VALUES (?, ?, ?, ?)
                """,
                (collection_id, json.dumps({"days": days, "targets": targets}), now, now),
            )
        progress, streak = _get_goal_progress(conn, collection_id)
    return jsonify(
        {
            "collection_id": collection_id,
            "days": days,
            "targets": targets,
            "progress": progress,
            "streak_days": streak,
            "updated_at": now,
        }
    )


@app.route("/api/goals/<int:collection_id>", methods=["DELETE"])
def delete_goals(collection_id):
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        conn.execute("DELETE FROM study_goals WHERE collection_id = ?", (collection_id,))
    return jsonify({"ok": True})


@app.route("/api/collections/<int:collection_id>/migrate", methods=["POST"])
def migrate_cards(collection_id):
    data, error = _get_request_json()
    if error:
        return error
    target_id, error = _parse_int(
        data.get("target_collection_id"), "target_collection_id", required=True, min_value=1
    )
    if error:
        return error
    if target_id == collection_id:
        return api_error("A coleção de destino deve ser diferente.", 400, code="invalid_target")
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção de origem não encontrada.", 404, code="collection_not_found")
        if not _collection_exists(conn, target_id):
            return api_error("Coleção de destino não encontrada.", 404, code="collection_not_found")
        cur = conn.execute(
            "UPDATE cards SET collection_id = ? WHERE collection_id = ?",
            (target_id, collection_id),
        )
    return jsonify({"ok": True, "moved": cur.rowcount})


@app.route("/api/import", methods=["POST"])
def import_cards():
    data, error = _get_request_json()
    if error:
        return error
    cards = data.get("cards") or []
    if not isinstance(cards, list):
        return api_error("cards inválido.", 400, code="invalid_cards")
    collection_id, error = _parse_int(
        data.get("collection_id"), "collection_id", required=True, min_value=1
    )
    if error:
        return error

    created_at = datetime.now(timezone.utc).isoformat()
    saved = []
    with get_db() as conn:
        if not _collection_exists(conn, collection_id):
            return api_error("Coleção não encontrada.", 404, code="collection_not_found")
        for card in cards:
            question = str(card.get("question") or "").strip()
            answer = str(card.get("answer") or "").strip()
            if not question or not answer:
                continue
            cur = conn.execute(
                """
                INSERT INTO cards (collection_id, question, answer, created_at)
                VALUES (?, ?, ?, ?)
                """,
                (
                    collection_id,
                    question,
                    answer,
                    created_at,
                ),
            )
            saved.append(
                {
                    "id": cur.lastrowid,
                    "collection_id": collection_id,
                    "question": question,
                    "answer": answer,
                    "created_at": created_at,
                }
            )
    return jsonify({"created": saved, "count": len(saved)})


init_db()

if __name__ == "__main__":
    app.run(debug=True, host="0.0.0.0", port=5000)
